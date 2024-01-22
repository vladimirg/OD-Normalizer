#!/usr/bin/env python3

import pandas as pd
from openpyxl import load_workbook
import os
from gooey import GooeyParser, Gooey 
from itertools import product

def parse_excluded_wells(excluded_str):
    """
    >>> parse_excluded_wells('')
    []
    >>> parse_excluded_wells('B3:C5')
    ['B3', 'B4', 'B5', 'C3', 'C4', 'C5']
    >>> parse_excluded_wells('G5,,F6')
    ['G5', 'F6']
    >>> parse_excluded_wells(' G5 ,B3:  C5, F6  ')
    ['G5', 'B3', 'B4', 'B5', 'C3', 'C4', 'C5', 'F6']
    >>> parse_excluded_wells('A10:A12')
    ['A10', 'A11', 'A12']
    """
    
    result = []
    
    parts = excluded_str.replace(" ", "").split(',')
    for p in parts:
        if ":" in p:
            start_well, stop_well = p.split(":")
            start_row, start_col = start_well[0], int(start_well[1:])
            stop_row, stop_col = stop_well[0], int(stop_well[1:])
            
            row_range = "ABCDEFGH"
            for row_label in row_range[row_range.index(start_row):
                                       row_range.index(stop_row)+1]:
                for col in range(start_col, stop_col+1):
                    result.append(f"{row_label}{col}")
        elif p:
            result.append(p)
    
    return result

def keep_only_first_unique(lst):
    """
    >>> keep_only_first_unique([])
    []
    >>> keep_only_first_unique(["A1", "A2"])
    ['A1', 'A2']
    >>> keep_only_first_unique(["A1", "A2", "A1", "A1", "A2", "A1"])
    ['A1', 'A2']
    """
    return [w for ix, w in enumerate(lst) if ix == lst.index(w)]

# NB: run with --ignore-gooey to force the CLI.
# NB: run using pythonw (instead of regular python) on the command line to
#     invoke this script as a GUI.
@Gooey(program_name="OD Normalizer")
def main():
    #### Parser set up
    parser = GooeyParser()
    parser.add_argument(
        "--in-file", type=str, widget="FileChooser",
        metavar="Input Excel file", required=True,
        help="The input Excel file from Sunrise. The ODs should be stored in Sheet1 within the workbook.")
    parser.add_argument("--target-od", type=float, metavar="Target OD", required=True)
    parser.add_argument(
        "--blank-wells", metavar="Wells to use as blanks", type=str,
        help="The mean of the blank wells will be substracted from the OD of the "+
        "Source wells. Blank wells are considered as 'excluded' wells. "+
        "Leave empty to use the nominal OD. Example: A4,B5:H7,C10")
    parser.add_argument(
        "--final-volume", type=int, default=200,
        metavar="Final volume (uL)", help="The final volume in the Target plate.")
    parser.add_argument(
        "--min-pipette", type=int, default=5,
        metavar="Minimum pipetting volume (uL)")
    parser.add_argument(
        "--max-pipette", type=int, default=197,
        metavar="Maximum pipetting volume (uL)"),
    parser.add_argument(
        "--exclude-wells", metavar="Wells to exclude", type=str,
        help="Excluded wells will be ignored in OD calculationgs and not "+
        "receive anything from the Source. Example: A4,B5:H7,C10")
    parser.add_argument(
        "--no-ddw-in-excluded", action="store_true",
        metavar="Keep excluded wells empty",
        help="By default, dispense the final volume (or the maximum pippetting "+
        "volume) of DDW into excluded wells. If this is checked, keep them empty.")
    parser.add_argument(
        "--source_is_target", action="store_true",
        metavar="Dilute into the Source plate",
        help="Make the dilutions in the Source plate (no Target plate needed). "+
        "If you specify this, you MUST also specify a column or a row offset.")
    parser.add_argument(
        "--row-offset", type=int, default=0,
        metavar="Row offset in Target",
        help="Positive offsets: A->H, negative: H->A. Wraps around!")
    parser.add_argument(
        "--col-offset", type=int, default=0,
        metavar="Column offset in Target",
        help="Positive offsets: 1->12, negative: 12->1. Wraps around!")
    parser.add_argument(
        "--out-folder", type=str, widget="DirChooser",
        metavar="Output folder",
        help="The output folder where 'ddw.csv' and 'source.csv' will be saved.")
    
    #### Argument parsing and reporting
    args = parser.parse_args()
    
    row_offset = args.row_offset
    col_offset = args.col_offset
    in_file = args.in_file
    out_folder = args.out_folder
    target_od = args.target_od
    target_vol = args.final_volume
    min_pipette = args.min_pipette
    # NB: (target_vol - min_pipette) is useful here so that we can guarantee
    # that the target_vol will be reached exactly in certain edge cases.
    max_pipette = min(args.max_pipette, target_vol - min_pipette)
    
    try:
        # If this argument is empty, it will be parsed as None, not as an empty string:
        blank_wells = parse_excluded_wells(args.blank_wells if args.blank_wells is not None else '')
    except:
        from traceback import format_exc
        print("ERROR: 'Wells to use as blanks' argument is invalid, aborting. Traceback:")
        print(format_exc())
        return
    
    try:
        # If this argument is empty, it will be parsed as None, not as an empty string:
        excluded_wells = parse_excluded_wells(args.exclude_wells if args.exclude_wells is not None else '')
    except:
        from traceback import format_exc
        print("ERROR: 'Wells to exclude' argument is invalid, aborting. Traceback:")
        print(format_exc())
        return
    
    if blank_wells:
        blank_wells = keep_only_first_unique(blank_wells)
        print(f"Using the following wells as blanks: {', '.join(blank_wells)}")
    
    excluded_wells += blank_wells
    if excluded_wells:
        excluded_wells = keep_only_first_unique(excluded_wells)
        print(f"Excluding wells: {', '.join(excluded_wells)}")
    
    #### Loading the input file
    wb = load_workbook(filename=in_file)
    sheet = wb["Sheet1"]
    
    od_excel_type = None
    for row_ix, row in enumerate(sheet.iter_rows()):
        cell_value = str(row[0].value)
        if "Rawdata" in cell_value:
            od_excel_type = "Sunrise"
        elif "<>" in str(row[0].value): # From F200
            od_excel_type = "F200" # Also works for the F50
        
        if od_excel_type:
            break
    
    if od_excel_type is None:
        print("ERROR: 'Sheet1' is malformed - has neither a 'Rawdata' or '<>'" +
              " prefix. Can't read ODs, aborting.")
        return
    
    # TODO: we don't handle the case where the table is partial and the indices
    # are supplied along with it.
    # NB: we allow for empty ODs - these are marked with -1 and are expected
    # to be ignored as part of 'excluded wells'.
    df = pd.read_excel(
        in_file,
        skiprows=row_ix+1 if od_excel_type == "Sunrise" else row_ix,
        nrows=8,
        usecols="B:M",
    ).set_index(pd.Index(list("ABCDEFGH"))).fillna(-1)
    
    #### Normalizing the OD relative to blank wells
    if blank_wells:
        blank_od = 0
        for well in blank_wells:
            row, col = well[0], int(well[1:])
            blank_od += df.loc[row, col]
        blank_od = blank_od/len(blank_wells)
        
        print(f"The blank OD is: {blank_od}")
        
        df = df - blank_od
    
    #### Calculating the Source and DDW volumes
    source_df = (target_od * target_vol / df).round().astype(int)
    ddw_df = pd.DataFrame(index=df.index, columns=df.columns)
    
    df_labels = list(product(df.index, df.columns))
    for row_letter, col_num in df_labels:
        well_name = f"{row_letter}{col_num}"
        ddw = None
        new_source_vol = None
        
        # TODO: we can't gurantee that the final volume will match the target
        # volume for edge cases - this may require double pipetting, depending
        # on the permissible min/max and target volume, which is too much of a 
        # hassle.
        if well_name in excluded_wells:
            new_source_vol = 0
            if args.no_ddw_in_excluded:
                ddw = 0
            else:
                ddw = args.max_pipette
        else:
            curr_source_vol = source_df.loc[row_letter, col_num]
            new_source_vol = curr_source_vol
            ddw = max(min(target_vol - curr_source_vol, max_pipette), min_pipette)
            well_ix = f"{'ABCDEFGH'.index(row_letter)+1+(col_num-1)*8}"
            if curr_source_vol < min_pipette:
                new_source_vol = min_pipette
                ddw = min(target_vol - new_source_vol, max_pipette)
                print(f"{well_name} ({well_ix}) is too concentrated (OD={df.loc[row_letter,col_num]}), "+
                      f"defaulting to taking the minimum pipetting volume ({min_pipette} uL). " +
                      f"Expected OD is {min_pipette*df.loc[row_letter,col_num]/(new_source_vol+ddw):.3}.")
            elif curr_source_vol > max_pipette:
                new_source_vol = args.max_pipette
                ddw = 0
                print(f"{well_name} ({well_ix}) is too diluted (OD={df.loc[row_letter,col_num]}), "+
                      f"defaulting to taking the maximum pipetting volume ({max_pipette} uL) and no DDW. " +
                      f"Expected OD is {max_pipette*df.loc[row_letter,col_num]/(new_source_vol+ddw):.3}.")
        
        assert ddw is not None
        ddw_df.loc[row_letter, col_num] = ddw
        assert new_source_vol is not None
        source_df.loc[row_letter, col_num] = new_source_vol
    
    #### Transpose the rows and columns in ddw_df
    if row_offset != 0:
        old_index = ddw_df.index
        ddw_df = pd.concat([
            ddw_df.iloc[-row_offset:],
            ddw_df.iloc[:-row_offset]
        ])
        ddw_df.set_index(old_index, inplace=True)
    
    if col_offset != 0:
        cols = ddw_df.columns
        ddw_df = ddw_df[list(cols[-col_offset:]) + list(cols[:-col_offset])]
        ddw_df.columns = cols
    
    #### Writing the output files
    ddw_fname = "ddw.csv"
    source_fname = "source.csv"
    
    # The ddw-to-target is NOT a worklist, but instead a CSV file from which
    # a special script (DDW_to_Target_variable_volume.exd) sent by Neotec can
    # read the data.
    # The format is a CSV where the first column is the row label ("Tip" and
    # "Volume", but can be anything), and the subsequent columns are the data -
    # the tip identifiers (1-8) on the first row and then the volumes on the
    # second row.
    with open(os.path.join(out_folder, ddw_fname), "w") as ddw_file:
        # NB: it's important for the entire plate to be included, as the wells
        # are processed in order, and the "Tip" column is just for humans and
        # is ignored by the robot. And while it may be possible to tell the
        # robot how many rows are required, in the existing way we're throwing
        # out at most 7 tips per plate, which isn't that bad.
        ddw_file.write(f"Tip,{','.join(str(i) for i in list(range(1, 9))*12)}\n")
        ddw_file.write(f"Volume,{','.join(','.join(ddw_df[i].astype(str)) for i in ddw_df)}\n")
    
    # The source-to-target is a worklist (CSV) with the following format:
    # Output format by columns:
    # 1 - source name
    # 2 - source position
    # 3 - target name
    # 4 - target position
    # 5 - volume
    with open(os.path.join(out_folder, source_fname), "w") as source_file:
        for col_ix, column in enumerate(source_df):
            series = source_df.loc[:, column]
            for vol_ix, vol in enumerate(series):
                # NB: this will not prevent the robot from taking a tip, but it
                # will force it to use the tips it took.
                if vol == 0:
                    continue
                
                source_pos = vol_ix+1+col_ix*8
                target_pos = (vol_ix+1+row_offset-1)%8+1 + col_ix*8
                target_pos = (target_pos + col_offset*8-1) % 96+1 
                target_label = "Target" if not args.source_is_target else "Source"
                source_file.write(f"Source,{source_pos},{target_label},{target_pos},{vol}\n")
                
    print("Done!\n")
    

if __name__ == "__main__":
    main()